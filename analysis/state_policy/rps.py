# %%
# SETUP
import pandas as pd
import numpy as np
from tqdm import tqdm
import dropbox
import io
import sys, os
import re
import openpyxl

parent_dir = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
sys.path.insert(0, parent_dir)
from credentials import DB_ACCESS_TOKEN

# options
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', 100)

# dropbox access
dbx = dropbox.Dropbox(DB_ACCESS_TOKEN)
rps_path = '/Net Zero Cluster/data/State Policy/rps/data_raw/'


# -------------------------------------
# %% ============ 2009 - 2012 ============
# -------------------------------------
# FROM JOE-ALDY PROVIDED RPS ARCHIVES: 2009-2012
# _, j = dbx.files_download(rps_path + 'aldy_state_rps_laws_and_data.xlsx')
# workbook = openpyxl.load_workbook(io.BytesIO(j.content), data_only=False)
# sheet_names = workbook.sheetnames
# sheet = workbook["Dec '12"]

def get_column_names(sheet, header_rows, data_start_row):
    """
    Generate column names based on the header rows.
    """
    headers = []
    for row in sheet.iter_rows(min_row=data_start_row - header_rows, 
                               max_row=data_start_row - 1, values_only=True):
        headers += [list(row)]
    headers[0] = pd.Series(headers[0]).ffill().to_list() # fill across merged columns
    # drop last column if last row is None
    headers = np.array(headers)
    if headers[-1, -1] is None:
        headers = headers[:, :-1]
    column_names = [' '.join(map(str,filter(None, col))).strip() for col in zip(*headers)]
    # column_names.insert(0, 'notes_comment')  # Insert the column for comments at the beginning
    column_names = [
        re.sub(r'\s+', '_',
            re.sub(r'\(.*?\)', '', c).strip())
        .strip().lower()
        for c in column_names
    ]
    # column_names = [
    #     re.sub('rps_start_year', 'start_year', c)
    #     for c in column_names]
    return column_names + ['all_comments']


def read_sheet(sheet, data_start_row):
    """
    Read specified columns from a sheet starting from a given row, including comments.
    """
    data = []
    for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
        data.append(list(row))
    # Check if the last column is empty (all None)
    if all(row[-1] is None for row in data):
        # Drop the last column
        data = [row[:-1] for row in data]
    return data

def handle_merged_cells(sheet, data, data_start_row):
    """
    Modify data to handle merged cells in specified columns.
    """
    for merged_cell in sheet.merged_cells.ranges:
        top_left_cell_value = sheet.cell(merged_cell.min_row, merged_cell.min_col).value
        for row in range(merged_cell.min_row, merged_cell.max_row + 1):
            if row >= data_start_row:  # Ensure we're in the data rows
                for col in range(merged_cell.min_col, merged_cell.max_col + 1):
                    data[row - data_start_row][col - 1] = top_left_cell_value
    return data

def aggregate_comments(sheet, data_start_row, column_names):
    """
    Aggregate comments from each cell, prefixed by the column name.
    """
    aggregated_comments = []
    for row in sheet.iter_rows(min_row=data_start_row, values_only=False):
        comments_in_row = []
        for col, cell in enumerate(row, start=1):
            comment = cell.comment
            if comment is not None:
                column_name = column_names[col - 1]
                comments_in_row.append(f"{column_name}: {comment.text.strip()}")
        aggregated_comment = " | ".join(comments_in_row)
        aggregated_comments.append(aggregated_comment)
    return aggregated_comments

def process_workbook(file_content, sheet_config):
    """
    Process each sheet in the workbook according to the provided configuration.
    """
    workbook = openpyxl.load_workbook(file_content, data_only=True)
    all_data = []
    for sheet_name, config in sheet_config.items():
        sheet = workbook[sheet_name]
        data_start_row = config['data_start_row']
        header_rows = config['header_rows']

        column_names = get_column_names(sheet, header_rows, data_start_row)
        data = read_sheet(sheet, data_start_row)
        data = handle_merged_cells(sheet, data, data_start_row)

        # Aggregate comments and append them to each row
        comments = aggregate_comments(sheet, data_start_row, column_names)
        for i, row in enumerate(data):
            row.append(comments[i])

        # Convert to DataFrame and set column names
        df = pd.DataFrame(data, columns=column_names)
        df['date'] = pd.to_datetime(sheet_name, format="%b '%y")
        all_data.append(df)

    # Concatenate all dataframes
    return pd.concat(all_data, ignore_index=True)

# Configuration for each sheet
sheet_config = {
    "Dec '12":{'data_start_row': 10, 'header_rows': 2}, 
    "Dec '11":{'data_start_row': 10, 'header_rows': 2}, 
    "Dec '10":{'data_start_row': 10, 'header_rows': 2}, 
    "Dec '09":{'data_start_row': 10, 'header_rows': 2}, 
    "Apr '09":{'data_start_row': 3, 'header_rows': 2}
}

# Use the function with your Excel file
_, file = dbx.files_download(rps_path + 'aldy_state_rps_laws_and_data.xlsx')
file_content = io.BytesIO(file.content)
output_df = process_workbook(file_content, sheet_config)

# # upload to dropbox
id_cols = [
    'date', 'state', 'memo_notes_and_updates', 'all_comments', 'rps_type', 'tier']
id_cols += [col for col in output_df if col.startswith('eligible')]
output_df = output_df[id_cols + [col for col in output_df.columns if col not in id_cols]]
try:
    dbx.files_upload(
        output_df.to_csv(index=False).encode(), 
        rps_path + f'df_cln_dsire_{output_df.date.dt.year.min()}_{output_df.date.dt.year.max()}.csv',
        mode=dropbox.files.WriteMode.overwrite)
    print("File uploaded successfully")
except dropbox.exceptions.ApiError as err:
    print(f"API Error: {err}")












# %%
# -------------------------------------
# ============ 2015-2023 ============
# -------------------------------------
# FROM ARCHIVED DATA IN RPS ARCHIVES: 2015-2023
# get relevant archives

entries = dbx.files_list_folder(rps_path)
archives = [f.name for f in entries.entries if f.name.startswith('dsire-')]

# For each archive, download, transform, and stack
# Use the Dropbox API to stream the file's content
cols_keep = ['program_id', 'state_id', 'code', 'name', 'websiteurl', 'summary']
def get_archive(rps_path, archive, cols_keep=cols_keep):
    # read in files
    _, p = dbx.files_download(rps_path + archive + '/program.csv')
    _, t = dbx.files_download(rps_path + archive + '/program_detail.csv')
    _, s = dbx.files_download(rps_path + archive + '/state.csv')
    p = pd.read_csv(io.BytesIO(p.content))
    t = pd.read_csv(io.BytesIO(t.content))
    s = pd.read_csv(io.BytesIO(s.content))

    # clean files
    t['label'] = t.label.str.lower().str.replace(r'\s|//', '_', regex=True)
    t = (t
         .drop(columns=['id', 'display_order', 'template_id'],)
         .loc[~t.label.isin(['alternative_compliance_payment', 'compliance_multipliers', 
              'credit_trading/tracking_system'])])
    p = (p
        .rename(columns={'id':'program_id'})
        .loc[(p.program_type_id == 38), cols_keep])
    p['summary'] = (p.summary
        .str.replace(r'<([^>]*)>', '', regex=True)
        .str.replace(r'\\[a-zA-Z]', '', regex=True)
        .str.replace(r'&#\d+', '', regex=True))
    s.rename(columns={'id':'state_id', 'name':'state_name'}, inplace=True)
    # merge files
    pt = pd.merge(left=p, right=t, how='left', on='program_id')
    pt = pt.pivot(
        index=cols_keep, values='value', columns='label').reset_index()
    pts = pd.merge(left=pt, right=s, how='inner', on='state_id')
    pts = pts.loc[(pts.is_territory == 0) | (pts.abbreviation == 'DC')]
    pts.drop(columns='is_territory', inplace=True)
    return pts

dfs = pd.DataFrame()
for archive in tqdm(archives):
    df = get_archive(rps_path, archive)
    df['date'] = pd.to_datetime(archive[6:])
    dfs = pd.concat([df, dfs], ignore_index=True)
id_cols = ['date', 'program_id', 'state_id', 'abbreviation', 'state_name']
dfs = dfs[id_cols + [col for col in dfs.columns if col not in id_cols]]

# upload to dropbox
try:
    dbx.files_upload(
        dfs.to_csv(index=False).encode(), 
        rps_path + f'df_cln_dsire_{dfs.date.dt.year.min()}_{dfs.date.dt.year.max()}.csv',
        mode=dropbox.files.WriteMode.overwrite)
    print("File uploaded successfully")
except dropbox.exceptions.ApiError as err:
    print(f"API Error: {err}")



# %%
# -------------------------------------
# ============ Appendix ============
# -------------------------------------
# APPENDIX - RPS USING NCSL DATA
from bs4 import BeautifulSoup
import requests
from tqdm import tqdm
import re
url = 'https://www.ncsl.org/energy/state-renewable-portfolio-standards-and-goals/maptype/'
req = requests.get(url)
soup = BeautifulSoup(req.content, 'html.parser')

# %%
# Initialize an empty list to store your data
data = []

# Loop through each state
for state_tag in tqdm(soup.find_all('h3')):
    state_name = state_tag.get_text()
    if state_name == 'Puerto Rico':
        break
    ul = state_tag.find_next_sibling('ul')
    lis = ul.findAll('li')
    for li in lis:
        # Extract each piece of information from the <li> tag
        text = li.get_text()
        label, value = text.split(':', 1)
        data += [{
            'state': state_name,
            'label': label.strip(),
            'value': value.strip()
        }]

# %%
rps = pd.DataFrame(data)
rps = rps.pivot(index='state', columns='label', values='value').reset_index()
rps.columns = rps.columns.str.lower().str.replace(' ', '_')

# Clean data
# 1. Create categorical columns for applicable sectors
sectors = ['investor-owned utility', 'municipal utilities', 'cooperative utilities', 'local government', 'retail supplier']
for sector in sectors:
    rps[sector] = rps.applicable_sectors.str.lower().str.contains(sector)
rps.drop(columns='applicable_sectors', inplace=True)
rps.to_csv('spd_rps_raw.csv', index=False)

# 2. Parse targets into long format
# cool, but not worth it
# rps['requirements'] = rps['requirement'].apply(lambda x: re.findall(r'(\d+)%\s*(.*?)\s*(by|from)\s*(.*?)\s*(\d+)', x))
# reqs = []
# for __, row in rps.iterrows():
#     for target, year in row['requirements']:
#         reqs.append({'state': row.state, 'target': target, 'year': year})
# reqs = pd.DataFrame(reqs)
# rpsl = pd.merge(left=rps.drop(columns='requirements'), 
#                right=reqs, how='outer', on=['state'])


# %%
rps
# %%
